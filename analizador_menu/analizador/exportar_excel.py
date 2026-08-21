"""Generacion del archivo de Excel con los resultados de la busqueda.

El libro tiene tres hojas:

* **Resultados** -- un renglon por item / restaurante / periodo. Las primeras
  columnas son las pedidas: #ID de restaurante, nombre, numero de item y
  cuantas ventas (unidades vendidas).
* **Resumen**    -- total de unidades e importe por item y periodo.
* **Archivos**   -- registro de los archivos leidos y lo que se dedujo de cada
  nombre (#restaurante y periodo).
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .busqueda import RegistroArchivo, ResultadoItem

RELLENO_ENCABEZADO = PatternFill("solid", fgColor="1F4E78")
FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF")

COLUMNAS_RESULTADOS = [
    ("#ID Restaurante", 16),
    ("Restaurante", 26),
    ("# Item", 10),
    ("Ventas (unidades)", 18),
    ("Periodo", 12),
    ("Fecha inicio", 14),
    ("Fecha fin", 14),
    ("Nombre del item", 24),
    ("Categoria", 22),
    ("Subcategoria", 24),
    ("Ventas ($)", 14),
    ("Precio de menu", 15),
    ("Costo total", 14),
    ("Utilidad", 14),
    ("Busqueda", 14),
    ("Encontrado", 12),
    ("Archivo", 26),
]

COLUMNAS_RESUMEN = [
    ("Busqueda", 14),
    ("# Item", 10),
    ("Nombre del item", 24),
    ("Periodo", 12),
    ("Restaurantes con venta", 22),
    ("Ventas (unidades)", 18),
    ("Ventas ($)", 16),
]

COLUMNAS_ARCHIVOS = [
    ("Archivo", 30),
    ("#ID Restaurante", 16),
    ("Restaurante", 26),
    ("Periodo", 12),
    ("Fecha inicio", 14),
    ("Fecha fin", 14),
    ("Items leidos", 14),
    ("Estado", 30),
]


def _escribir_encabezado(hoja, columnas: Sequence[tuple[str, int]]) -> None:
    for indice, (titulo, ancho) in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=indice, value=titulo)
        celda.fill = RELLENO_ENCABEZADO
        celda.font = FUENTE_ENCABEZADO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[get_column_letter(indice)].width = ancho
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}1"


def _formato_fecha(valor: date | None):
    return valor.strftime("%d/%m/%Y") if valor else ""


def _hoja_resultados(libro: Workbook, resultados: Sequence[ResultadoItem]) -> None:
    hoja = libro.active
    hoja.title = "Resultados"
    _escribir_encabezado(hoja, COLUMNAS_RESULTADOS)

    for fila, resultado in enumerate(resultados, start=2):
        valores = [
            resultado.restaurante_id,
            resultado.restaurante_nombre,
            resultado.item_numero if resultado.item_numero is not None else resultado.busqueda,
            resultado.unidades_vendidas,
            resultado.periodo,
            _formato_fecha(resultado.fecha_inicio),
            _formato_fecha(resultado.fecha_fin),
            resultado.item_nombre,
            resultado.categoria,
            resultado.subcategoria,
            round(resultado.ventas_importe, 2),
            round(resultado.precio_menu, 2),
            round(resultado.costo_total, 2),
            round(resultado.utilidad, 2),
            resultado.busqueda,
            "Si" if resultado.encontrado else "No",
            resultado.archivo,
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
        for columna in (11, 12, 13, 14):
            hoja.cell(row=fila, column=columna).number_format = "#,##0.00"

    hoja.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS_RESULTADOS))}{max(len(resultados) + 1, 1)}"


def _hoja_resumen(libro: Workbook, resultados: Sequence[ResultadoItem]) -> None:
    hoja = libro.create_sheet("Resumen")
    _escribir_encabezado(hoja, COLUMNAS_RESUMEN)

    acumulado: dict[tuple, dict] = defaultdict(
        lambda: {"nombre": "", "restaurantes": set(), "unidades": 0, "importe": 0.0}
    )
    for resultado in resultados:
        clave = (resultado.busqueda, resultado.item_numero, resultado.periodo)
        registro = acumulado[clave]
        if resultado.item_nombre and not registro["nombre"]:
            registro["nombre"] = resultado.item_nombre
        if resultado.unidades_vendidas:
            registro["restaurantes"].add(resultado.restaurante_id)
        registro["unidades"] += resultado.unidades_vendidas
        registro["importe"] += resultado.ventas_importe

    fila = 2
    for (busqueda, numero, periodo), registro in sorted(
        acumulado.items(), key=lambda par: (str(par[0][0]), str(par[0][1]), str(par[0][2]))
    ):
        valores = [
            busqueda,
            numero if numero is not None else "",
            registro["nombre"],
            periodo,
            len(registro["restaurantes"]),
            registro["unidades"],
            round(registro["importe"], 2),
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
        hoja.cell(row=fila, column=7).number_format = "#,##0.00"
        fila += 1


def _hoja_archivos(
    libro: Workbook,
    registros: Sequence[RegistroArchivo],
    resultados: Sequence[ResultadoItem],
) -> None:
    hoja = libro.create_sheet("Archivos")
    _escribir_encabezado(hoja, COLUMNAS_ARCHIVOS)

    nombres = {
        resultado.restaurante_id: resultado.restaurante_nombre
        for resultado in resultados
        if resultado.restaurante_nombre
    }

    for fila, registro in enumerate(registros, start=2):
        if registro.reporte is None:
            estado = f"ERROR: {registro.error}"
        elif not registro.reporte.items:
            estado = "SIN ITEMS (revisar el formato del archivo)"
        else:
            estado = "OK"
        valores = [
            registro.info.nombre_archivo,
            registro.restaurante_id,
            nombres.get(registro.restaurante_id, ""),
            registro.periodo,
            _formato_fecha(registro.fecha_inicio),
            _formato_fecha(registro.fecha_fin),
            len(registro.reporte.items) if registro.reporte else 0,
            estado,
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)


def exportar_resultados(
    resultados: Sequence[ResultadoItem],
    ruta_salida: str | Path,
    registros: Sequence[RegistroArchivo] = (),
) -> Path:
    """Escribe el Excel de resultados y devuelve la ruta final."""
    ruta_salida = Path(ruta_salida)
    if ruta_salida.suffix.lower() != ".xlsx":
        ruta_salida = ruta_salida.with_suffix(".xlsx")
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    # Si el archivo esta abierto en Excel no se puede sobrescribir: se versiona.
    if ruta_salida.exists():
        try:
            with open(ruta_salida, "ab"):
                pass
        except OSError:
            marca = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_salida = ruta_salida.with_name(f"{ruta_salida.stem}_{marca}.xlsx")

    libro = Workbook()
    _hoja_resultados(libro, resultados)
    _hoja_resumen(libro, resultados)
    if registros:
        _hoja_archivos(libro, registros, resultados)
    libro.save(ruta_salida)
    return ruta_salida

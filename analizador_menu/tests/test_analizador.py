"""Pruebas del analizador.  Ejecutar con:  python -m unittest discover tests"""

from __future__ import annotations

import re
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from analizador.busqueda import buscar_items, cargar_reportes, normalizar_busquedas
from analizador.diccionario import DiccionarioRestaurantes
from analizador.exportar_excel import exportar_resultados
from analizador.nombre_archivo import interpretar_nombre, listar_archivos
from analizador.parser_reporte import leer_reporte

REPORTE_EJEMPLO = RAIZ / "datos" / "6001 (3-6).txt"
DICCIONARIO = RAIZ / "Diccionario_restaurantes.xlsx"


class PruebaNombreArchivo(unittest.TestCase):
    def test_formato_estandar(self):
        info = interpretar_nombre("6001 (3-6).txt")
        self.assertEqual(info.restaurante_id, "6001")
        self.assertEqual(info.periodo_normalizado, "3-6")

    def test_periodo_de_dos_digitos(self):
        info = interpretar_nombre("/reportes/6002 (10-13).txt")
        self.assertEqual(info.restaurante_id, "6002")
        self.assertEqual((info.dia_inicio, info.dia_fin), (10, 13))

    def test_variantes_de_nombre(self):
        for nombre, periodo in [
            ("6013_3-6.txt", "3-6"),
            ("6013-10-13.txt", "10-13"),
            ("6013 (3 al 6).txt", "3-6"),
            ("6013 [Ago 10-13].txt", "10-13"),
        ]:
            with self.subTest(nombre=nombre):
                info = interpretar_nombre(nombre)
                self.assertEqual(info.restaurante_id, "6013")
                self.assertEqual(info.periodo_normalizado, periodo)

    def test_sin_periodo(self):
        info = interpretar_nombre("6021.txt")
        self.assertEqual(info.restaurante_id, "6021")
        self.assertEqual(info.periodo_normalizado, "")


class PruebaParser(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.reporte = leer_reporte(REPORTE_EJEMPLO)

    def test_encabezado(self):
        self.assertEqual(self.reporte.restaurante_id, "6001")
        self.assertIn("BOSTONS", self.reporte.restaurante_encabezado)
        self.assertEqual(self.reporte.fecha_inicio, date(2026, 8, 3))
        self.assertEqual(self.reporte.fecha_fin, date(2026, 8, 6))

    def test_columnas_de_un_item(self):
        (item,) = self.reporte.buscar_por_numero(37014)
        self.assertEqual(item.nombre, "PLATO DE EQUIPO")
        self.assertEqual(item.categoria, "1--STARTERS")
        self.assertEqual(item.subcategoria, "1--TEAM PLATTER")
        self.assertAlmostEqual(item.precio_menu, 399.00)
        self.assertEqual(item.unidades_vendidas, 98)
        self.assertAlmostEqual(item.ventas, 33865.87)
        self.assertAlmostEqual(item.costo_unitario_prom, 130.2717)
        self.assertAlmostEqual(item.costo_total, 12766.63)
        self.assertAlmostEqual(item.pct_costo, 37.6976)
        self.assertAlmostEqual(item.utilidad, 21099.24)

    def test_nombre_con_coma(self):
        (item,) = self.reporte.buscar_por_numero(37202)
        self.assertEqual(item.nombre, "ENS, DE LA CASA<")
        self.assertEqual(item.unidades_vendidas, 2)

    def test_todas_las_paginas_se_leen(self):
        """El reporte esta paginado: deben venir items de la primera y la ultima."""
        self.assertTrue(self.reporte.buscar_por_numero(37014))  # pagina 1
        self.assertTrue(self.reporte.buscar_por_numero(1998))   # ultimas paginas
        self.assertGreater(len(self.reporte.items), 700)

    def test_categorias_de_dos_digitos(self):
        """'  10--OPEN LIQUOR' es subcategoria, no categoria."""
        (item,) = self.reporte.buscar_por_numero(1998)
        self.assertEqual(item.categoria, "20--APPROVED PROMO")
        self.assertEqual(item.subcategoria, "5--LSM PROMOS")
        self.assertNotIn("10--OPEN LIQUOR", {i.categoria for i in self.reporte.items})

    def test_ningun_item_sin_categoria(self):
        self.assertFalse([i for i in self.reporte.items if not i.categoria])

    def test_no_se_leen_subtotales(self):
        nombres = {i.nombre.upper() for i in self.reporte.items}
        self.assertFalse({n for n in nombres if n.startswith(("SUBTOTAL", "TOTAL"))})

    def test_totales_cuadran_con_el_reporte(self):
        """La suma de #Sold debe coincidir con el renglon 'Category Totals'."""
        texto = REPORTE_EJEMPLO.read_text(encoding="utf-8")
        renglon = next(l for l in texto.splitlines() if l.startswith("Category Totals"))
        total_reporte = float(renglon[34:45].strip())
        total_leido = sum(i.num_vendidos for i in self.reporte.items)
        self.assertAlmostEqual(total_leido, total_reporte, places=2)

    def test_subtotales_por_categoria_cuadran(self):
        """Cada 'Subtotal <categoria>' debe coincidir con la suma de sus items."""
        texto = REPORTE_EJEMPLO.read_text(encoding="utf-8")
        patron = re.compile(r"^ Subtotal\s+(?P<cat>\d+--\S.*?)\s{2,}")
        revisados = 0
        for linea in texto.splitlines():
            coincidencia = patron.match(linea.rstrip("\r"))
            if not coincidencia:
                continue
            # El reporte recorta la etiqueta del subtotal, por eso se compara por prefijo.
            etiqueta = coincidencia.group("cat").strip()
            esperado = float(linea[34:45].strip())
            obtenido = sum(
                i.num_vendidos for i in self.reporte.items if i.categoria.startswith(etiqueta)
            )
            self.assertAlmostEqual(obtenido, esperado, places=2, msg=f"categoria {etiqueta}")
            revisados += 1
        self.assertGreaterEqual(revisados, 10)

    def test_busqueda_por_nombre(self):
        encontrados = self.reporte.buscar_por_nombre("plato de equipo")
        self.assertEqual([i.inv for i in encontrados], [37014])


class PruebaDiccionario(unittest.TestCase):
    def test_carga_desde_excel(self):
        diccionario = DiccionarioRestaurantes.desde_excel(DICCIONARIO)
        self.assertGreaterEqual(len(diccionario), 20)
        self.assertEqual(diccionario.nombre("6001"), "Merida")
        self.assertEqual(diccionario.nombre("6013"), "Playa del Carmen")
        self.assertEqual(diccionario.nombre("6008"), "Pensiones")  # "6008 -Pensiones"
        self.assertEqual(diccionario.nombre("9999", "?"), "?")

    def test_localizar_en_carpeta(self):
        self.assertIsNotNone(DiccionarioRestaurantes.localizar(RAIZ))


class PruebaBusqueda(unittest.TestCase):
    def test_normalizar_entradas(self):
        self.assertEqual(normalizar_busquedas(["37014, 37021"]), ["37014", "37021"])
        self.assertEqual(normalizar_busquedas(["37014-37016"]), ["37014", "37015", "37016"])
        self.assertEqual(normalizar_busquedas(["37014, 37014"]), ["37014"])
        self.assertEqual(normalizar_busquedas([" alitas "]), ["alitas"])

    def test_busqueda_en_carpeta_con_varios_periodos(self):
        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            original = REPORTE_EJEMPLO.read_text(encoding="utf-8")
            (carpeta / "6001 (3-6).txt").write_text(original, encoding="utf-8")
            (carpeta / "6002 (10-13).txt").write_text(
                original.replace("#6001", "#6002").replace("8/3/26 to 8/6/26", "8/10/26 to 8/13/26"),
                encoding="utf-8",
            )
            self.assertEqual(len(listar_archivos(carpeta)), 2)

            registros = cargar_reportes(carpeta)
            diccionario = DiccionarioRestaurantes.desde_excel(DICCIONARIO)
            resultados = buscar_items(registros, ["37014", "99999"], diccionario)

            # La clave del periodo son las fechas que vienen DENTRO del reporte.
            ventas = {
                (r.restaurante_id, r.fecha_inicio): r
                for r in resultados
                if r.item_numero == 37014
            }
            primero = ventas[("6001", date(2026, 8, 3))]
            self.assertEqual(primero.unidades_vendidas, 98)
            self.assertEqual(primero.restaurante_nombre, "Merida")
            self.assertEqual(primero.fecha_fin, date(2026, 8, 6))
            segundo = ventas[("6002", date(2026, 8, 10))]
            self.assertEqual(segundo.restaurante_nombre, "Altabrisa")
            self.assertEqual(segundo.fecha_fin, date(2026, 8, 13))

            faltantes = [r for r in resultados if r.busqueda == "99999"]
            self.assertEqual(len(faltantes), 2)
            self.assertTrue(all(not r.encontrado and r.unidades_vendidas == 0 for r in faltantes))

    def test_las_fechas_vienen_del_reporte_no_del_nombre(self):
        """Si el nombre del archivo miente, mandan las fechas del reporte."""
        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            # El nombre dice 20-23, pero adentro el reporte es del 3 al 6.
            (carpeta / "6001 (20-23).txt").write_text(
                REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
            )
            registros = cargar_reportes(carpeta)
            (resultado,) = buscar_items(registros, ["37014"])
            self.assertEqual(resultado.fecha_inicio, date(2026, 8, 3))
            self.assertEqual(resultado.fecha_fin, date(2026, 8, 6))
            self.assertFalse(hasattr(resultado, "periodo"))
            self.assertIn("20-23", registros[0].aviso_periodo)

    def test_sin_aviso_cuando_el_nombre_coincide(self):
        registros = cargar_reportes(REPORTE_EJEMPLO.parent)
        self.assertEqual(registros[0].aviso_periodo, "")

    def test_excluir_no_encontrados(self):
        registros = cargar_reportes(REPORTE_EJEMPLO.parent)
        resultados = buscar_items(registros, ["99999"], incluir_no_encontrados=False)
        self.assertEqual(resultados, [])


class PruebaExportacion(unittest.TestCase):
    def test_genera_las_tres_hojas(self):
        from openpyxl import load_workbook

        registros = cargar_reportes(REPORTE_EJEMPLO.parent)
        resultados = buscar_items(registros, ["37014"])
        with tempfile.TemporaryDirectory() as tmp:
            ruta = exportar_resultados(resultados, Path(tmp) / "salida", registros)
            self.assertTrue(ruta.exists())
            libro = load_workbook(ruta)
            self.assertEqual(libro.sheetnames, ["Resultados", "Resumen", "Archivos"])
            hoja = libro["Resultados"]
            encabezados = [c.value for c in hoja[1]]
            self.assertEqual(
                encabezados[:6],
                [
                    "#ID Restaurante",
                    "Restaurante",
                    "# Item",
                    "Ventas (unidades)",
                    "Fecha inicio",
                    "Fecha fin",
                ],
            )
            # Ya no hay columna de "Periodo" en ninguna hoja.
            for nombre_hoja in libro.sheetnames:
                self.assertNotIn("Periodo", [c.value for c in libro[nombre_hoja][1]])
            self.assertEqual(hoja.cell(row=2, column=3).value, 37014)
            self.assertEqual(hoja.cell(row=2, column=4).value, 98)
            self.assertEqual(hoja.cell(row=2, column=5).value, "03/08/2026")
            self.assertEqual(hoja.cell(row=2, column=6).value, "06/08/2026")


if __name__ == "__main__":
    unittest.main()

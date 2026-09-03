"""Pruebas del analizador.  Ejecutar con:  python -m unittest discover tests"""

from __future__ import annotations

import re
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent          # carpeta "sistema"
PROYECTO = RAIZ.parent                                  # carpeta del programa
sys.path.insert(0, str(RAIZ))

from analizador.busqueda import (
    buscar_items,
    cargar_reportes,
    detectar_problemas,
    normalizar_busquedas,
)
from analizador.configuracion import RESTAURANTES_ORACLE
from analizador.diccionario import DiccionarioRestaurantes
from analizador.exportar_excel import (
    MAXIMO_HOJAS_COMBINACION,
    _nombre_de_hoja,
    contar_combinaciones,
    exportar_resultados,
)
from analizador.nombre_archivo import interpretar_nombre, listar_archivos
from analizador.parser_reporte import leer_reporte

REPORTE_EJEMPLO = PROYECTO / "datos" / "6001 (3-6).txt"
REPORTE_PDF = PROYECTO / "datos" / "6012 (17-19).pdf"
DICCIONARIO = RAIZ / "Diccionario_restaurantes.xlsx"
FIXTURES = RAIZ / "tests" / "fixtures"
REPORTE_VACIO_PDF = FIXTURES / "6004_10_al_13.pdf"

# Un reporte que "se genero mal": la descarga corto el encabezado del
# restaurante (queda "NSO IMAGE 2017" en vez del nombre y el #ID) y no trae
# un solo producto, pero conserva el periodo y llega a "Category Totals" en
# cero.  Es el caso real que motivo esta validacion.
TEXTO_REPORTE_VACIO = """NSO IMAGE 2017                                                                                                             PAGE:   1
REPORT DATE: 09/02/2026                                                                                     REPORT TIME: 11:10:16.73
------------------------------------------------------------------------------------------------------------------------------------
                               Menu Item Food Cost Analysis - Item By Item (% based on Grand Totals)
                                                For the period:  8/10/26 to 8/13/26
 Menu Item                    Menu                                     % of        Avg.        Total            % Cont
    Inv# Name                Price      #Sold       #Del      Sales  Tot Sls    Unit Cost       Cost     Cost %   Margin     Profit
------------------------------------------------------------------------------------------------------------------------------------
Category Totals                          0.00       0.00       0.00    0.00%       0.0000       0.00    0.0000%    0.00%       0.00
"""


class PruebaNombreArchivo(unittest.TestCase):
    def test_formato_estandar(self):
        info = interpretar_nombre("6001 (3-6).txt")
        self.assertEqual(info.restaurante_id, "6001")
        self.assertEqual(info.periodo_normalizado, "3-6")

    def test_periodo_de_dos_digitos(self):
        info = interpretar_nombre("/reportes/6002 (10-13).txt")
        self.assertEqual(info.restaurante_id, "6002")
        self.assertEqual((info.dia_inicio, info.dia_fin), (10, 13))

    def test_variantes_de_nombre(self):
        for nombre, periodo in [
            ("6013_3-6.txt", "3-6"),
            ("6013-10-13.txt", "10-13"),
            ("6013 (3 al 6).txt", "3-6"),
            ("6013 [Ago 10-13].txt", "10-13"),
            ("6013_10_al_13.pdf", "10-13"),  # formato real que llega en PDF
            ("6013_24_al_27.pdf", "24-27"),
        ]:
            with self.subTest(nombre=nombre):
                info = interpretar_nombre(nombre)
                self.assertEqual(info.restaurante_id, "6013")
                self.assertEqual(info.periodo_normalizado, periodo)

    def test_sin_periodo(self):
        info = interpretar_nombre("6021.txt")
        self.assertEqual(info.restaurante_id, "6021")
        self.assertEqual(info.periodo_normalizado, "")


class PruebaParser(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.reporte = leer_reporte(REPORTE_EJEMPLO)

    def test_encabezado(self):
        self.assertEqual(self.reporte.restaurante_id, "6001")
        self.assertIn("BOSTONS", self.reporte.restaurante_encabezado)
        self.assertEqual(self.reporte.fecha_inicio, date(2026, 8, 3))
        self.assertEqual(self.reporte.fecha_fin, date(2026, 8, 6))

    def test_columnas_de_un_item(self):
        (item,) = self.reporte.buscar_por_numero(37014)
        self.assertEqual(item.nombre, "PLATO DE EQUIPO")
        self.assertEqual(item.categoria, "1--STARTERS")
        self.assertEqual(item.subcategoria, "1--TEAM PLATTER")
        self.assertAlmostEqual(item.precio_menu, 399.00)
        self.assertEqual(item.unidades_vendidas, 98)
        self.assertAlmostEqual(item.ventas, 33865.87)
        self.assertAlmostEqual(item.costo_unitario_prom, 130.2717)
        self.assertAlmostEqual(item.costo_total, 12766.63)
        self.assertAlmostEqual(item.pct_costo, 37.6976)
        self.assertAlmostEqual(item.utilidad, 21099.24)

    def test_importes_negativos_de_promociones(self):
        """Las promociones vienen con importe negativo en el propio reporte."""
        (item,) = self.reporte.buscar_por_numero(10258)
        self.assertEqual(item.categoria, "20--APPROVED PROMO")
        self.assertEqual(item.unidades_vendidas, 94)      # unidades positivas
        self.assertAlmostEqual(item.ventas, -10534.57)    # importe descontado
        self.assertAlmostEqual(item.utilidad, -10534.57)

    def test_signo_negativo_que_desborda_la_columna(self):
        """Si el '-' se sale de su columna, no se lee el importe como positivo."""
        from analizador.parser_reporte import _campo

        linea = (
            "   99999 ITEM DESBORDADO    99.00      10.00       0.00"
            "-12345678.90  -1.69%       0.0000       0.00    0.0000%   -2.54%  -1234.56"
        )
        self.assertEqual(_campo(linea, "ventas").strip(), "-12345678.90")
        self.assertEqual(_campo(linea, "pct_ventas_totales").strip(), "-1.69%")

    def test_nombre_con_coma(self):
        (item,) = self.reporte.buscar_por_numero(37202)
        self.assertEqual(item.nombre, "ENS, DE LA CASA<")
        self.assertEqual(item.unidades_vendidas, 2)

    def test_todas_las_paginas_se_leen(self):
        """El reporte esta paginado: deben venir items de la primera y la ultima."""
        self.assertTrue(self.reporte.buscar_por_numero(37014))  # pagina 1
        self.assertTrue(self.reporte.buscar_por_numero(1998))   # ultimas paginas
        self.assertGreater(len(self.reporte.items), 700)

    def test_categorias_de_dos_digitos(self):
        """'  10--OPEN LIQUOR' es subcategoria, no categoria."""
        (item,) = self.reporte.buscar_por_numero(1998)
        self.assertEqual(item.categoria, "20--APPROVED PROMO")
        self.assertEqual(item.subcategoria, "5--LSM PROMOS")
        self.assertNotIn("10--OPEN LIQUOR", {i.categoria for i in self.reporte.items})

    def test_ningun_item_sin_categoria(self):
        self.assertFalse([i for i in self.reporte.items if not i.categoria])

    def test_no_se_leen_subtotales(self):
        nombres = {i.nombre.upper() for i in self.reporte.items}
        self.assertFalse({n for n in nombres if n.startswith(("SUBTOTAL", "TOTAL"))})

    def test_totales_cuadran_con_el_reporte(self):
        """La suma de #Sold debe coincidir con el renglon 'Category Totals'."""
        texto = REPORTE_EJEMPLO.read_text(encoding="utf-8")
        renglon = next(l for l in texto.splitlines() if l.startswith("Category Totals"))
        total_reporte = float(renglon[34:45].strip())
        total_leido = sum(i.num_vendidos for i in self.reporte.items)
        self.assertAlmostEqual(total_leido, total_reporte, places=2)

    def test_subtotales_por_categoria_cuadran(self):
        """Cada 'Subtotal <categoria>' debe coincidir con la suma de sus items."""
        texto = REPORTE_EJEMPLO.read_text(encoding="utf-8")
        patron = re.compile(r"^ Subtotal\s+(?P<cat>\d+--\S.*?)\s{2,}")
        revisados = 0
        for linea in texto.splitlines():
            coincidencia = patron.match(linea.rstrip("\r"))
            if not coincidencia:
                continue
            # El reporte recorta la etiqueta del subtotal, por eso se compara por prefijo.
            etiqueta = coincidencia.group("cat").strip()
            esperado = float(linea[34:45].strip())
            obtenido = sum(
                i.num_vendidos for i in self.reporte.items if i.categoria.startswith(etiqueta)
            )
            self.assertAlmostEqual(obtenido, esperado, places=2, msg=f"categoria {etiqueta}")
            revisados += 1
        self.assertGreaterEqual(revisados, 10)

    def test_busqueda_por_nombre(self):
        encontrados = self.reporte.buscar_por_nombre("plato de equipo")
        self.assertEqual([i.inv for i in encontrados], [37014])


class PruebaParserPDF(unittest.TestCase):
    """El mismo reporte, entregado en PDF, se lee igual que el .txt."""

    @classmethod
    def setUpClass(cls):
        cls.reporte = leer_reporte(REPORTE_PDF)

    def test_encabezado_y_periodo(self):
        self.assertEqual(self.reporte.restaurante_id, "6012")
        self.assertIn("SANTA FE", self.reporte.restaurante_encabezado)
        self.assertEqual(self.reporte.fecha_inicio, date(2026, 8, 17))
        self.assertEqual(self.reporte.fecha_fin, date(2026, 8, 19))

    def test_columnas_de_un_item(self):
        (item,) = self.reporte.buscar_por_numero(37014)
        self.assertEqual(item.nombre, "PLATO DE EQUIPO")
        self.assertEqual(item.categoria, "1--STARTERS")
        self.assertEqual(item.subcategoria, "1--TEAM PLATTER")
        self.assertEqual(item.unidades_vendidas, 4)
        self.assertAlmostEqual(item.precio_menu, 414.00)
        self.assertAlmostEqual(item.ventas, 1433.27)
        self.assertAlmostEqual(item.costo_total, 509.58)

    def test_totales_cuadran_con_el_reporte(self):
        """La suma de #Sold debe coincidir con el renglon 'Category Totals'."""
        from analizador.parser_reporte import _leer_texto

        renglon = next(
            l for l in _leer_texto(REPORTE_PDF).splitlines() if l.startswith("Category Totals")
        )
        self.assertAlmostEqual(
            sum(i.num_vendidos for i in self.reporte.items),
            float(renglon[34:45].strip()),
            places=2,
        )

    def test_todas_las_paginas_se_leen(self):
        self.assertGreater(len(self.reporte.items), 300)
        self.assertFalse([i for i in self.reporte.items if not i.categoria])

    def test_no_se_leen_subtotales(self):
        nombres = {i.nombre.upper() for i in self.reporte.items}
        self.assertFalse({n for n in nombres if n.startswith(("SUBTOTAL", "TOTAL"))})


class PruebaCarpetaMixta(unittest.TestCase):
    """Se pueden mezclar .txt y .pdf en la misma carpeta."""

    def test_lee_los_dos_formatos(self):
        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            (carpeta / "6001 (3-6).txt").write_text(
                REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
            )
            (carpeta / "6012 (17-19).pdf").write_bytes(REPORTE_PDF.read_bytes())

            self.assertEqual(len(listar_archivos(carpeta)), 2)
            registros = cargar_reportes(carpeta)
            self.assertTrue(all(r.reporte for r in registros), [r.error for r in registros])

            resultados = buscar_items(registros, ["37014"])
            ventas = {(r.restaurante_id, r.fecha_inicio): r.unidades_vendidas for r in resultados}
            self.assertEqual(ventas[("6001", date(2026, 8, 3))], 98)
            self.assertEqual(ventas[("6012", date(2026, 8, 17))], 4)

    def test_un_pdf_sin_texto_no_tumba_el_programa(self):
        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            (carpeta / "6099 (1-2).pdf").write_bytes(b"%PDF-1.4 archivo roto")
            (registro,) = cargar_reportes(carpeta)
            self.assertIsNone(registro.reporte)
            self.assertTrue(registro.error)


class PruebaReporteVacio(unittest.TestCase):
    """Un reporte que se descargo mal: sin error, pero sin ningun producto."""

    def test_se_lee_sin_error_pero_sin_items(self):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / "6004_10_al_13.txt"
            ruta.write_text(TEXTO_REPORTE_VACIO, encoding="utf-8")
            reporte = leer_reporte(ruta)

            self.assertEqual(reporte.items, [])
            # El encabezado corrupto no trae ni nombre ni #ID de restaurante...
            self.assertIsNone(reporte.restaurante_id)
            self.assertIsNone(reporte.restaurante_encabezado)
            # ...pero el periodo si sobrevive, porque esa linea no se corto.
            self.assertEqual(reporte.fecha_inicio, date(2026, 8, 10))
            self.assertEqual(reporte.fecha_fin, date(2026, 8, 13))

    def test_pdf_real_vacio_de_ejemplo(self):
        """Regresion con un PDF real que llego asi de un restaurante."""
        reporte = leer_reporte(REPORTE_VACIO_PDF)
        self.assertEqual(reporte.items, [])
        self.assertEqual(reporte.fecha_inicio, date(2026, 8, 10))
        self.assertEqual(reporte.fecha_fin, date(2026, 8, 13))


class PruebaDetectarProblemas(unittest.TestCase):
    """Se avisa de los archivos vacios o rotos ANTES de buscar ningun item."""

    def _carpeta_de_prueba(self, tmp: str) -> Path:
        carpeta = Path(tmp)
        (carpeta / "6001 (3-6).txt").write_text(
            REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
        )
        # El restaurante 6004 "salio mal": sin un solo producto adentro.
        (carpeta / "6004_10_al_13.txt").write_text(TEXTO_REPORTE_VACIO, encoding="utf-8")
        # Y este ni siquiera se pudo abrir.
        (carpeta / "6099 (1-2).pdf").write_bytes(b"%PDF-1.4 archivo roto")
        return carpeta

    def test_esta_vacio_y_tiene_problema(self):
        with tempfile.TemporaryDirectory() as tmp:
            registros = cargar_reportes(self._carpeta_de_prueba(tmp))
            por_archivo = {r.info.nombre_archivo: r for r in registros}

            self.assertFalse(por_archivo["6001 (3-6).txt"].esta_vacio)
            self.assertFalse(por_archivo["6001 (3-6).txt"].tiene_problema)

            self.assertTrue(por_archivo["6004_10_al_13.txt"].esta_vacio)
            self.assertTrue(por_archivo["6004_10_al_13.txt"].tiene_problema)
            self.assertEqual(por_archivo["6004_10_al_13.txt"].error, "")  # no fue un error

            self.assertFalse(por_archivo["6099 (1-2).pdf"].esta_vacio)  # no hay reporte
            self.assertTrue(por_archivo["6099 (1-2).pdf"].tiene_problema)

    def test_detectar_problemas_clasifica_y_ordena(self):
        with tempfile.TemporaryDirectory() as tmp:
            registros = cargar_reportes(self._carpeta_de_prueba(tmp))
            problemas = detectar_problemas(registros)

            self.assertEqual(len(problemas), 2)
            motivos = {p.registro.info.nombre_archivo: p.motivo for p in problemas}
            self.assertEqual(motivos["6004_10_al_13.txt"], "vacio")
            self.assertEqual(motivos["6099 (1-2).pdf"], "error")
            # El restaurante 6001, que si esta bien, no aparece en la lista.
            self.assertNotIn("6001 (3-6).txt", motivos)
            # Ordenados por #ID de restaurante: primero el 6004, luego el que
            # no tiene ID identificable.
            self.assertEqual([p.registro.info.nombre_archivo for p in problemas][0], "6004_10_al_13.txt")

    def test_ningun_problema_cuando_todo_esta_bien(self):
        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            (carpeta / "6001 (3-6).txt").write_text(
                REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
            )
            registros = cargar_reportes(carpeta)
            self.assertEqual(detectar_problemas(registros), [])

    def test_el_restaurante_vacio_se_identifica_por_el_nombre_del_archivo(self):
        """Aunque el encabezado del reporte venga corrupto, el #ID sale del
        nombre del archivo, asi que el diccionario si puede dar el nombre."""
        with tempfile.TemporaryDirectory() as tmp:
            registros = cargar_reportes(self._carpeta_de_prueba(tmp))
            (vacio,) = [r for r in registros if r.esta_vacio]
            diccionario = DiccionarioRestaurantes.desde_excel(DICCIONARIO)
            self.assertEqual(vacio.restaurante_id, "6004")
            self.assertEqual(diccionario.nombre(vacio.restaurante_id), "Ciudad Del Carmen")

    def test_buscar_items_ignora_los_restaurantes_vacios(self):
        """El propio buscador ya no los cuenta como restaurantes con datos."""
        with tempfile.TemporaryDirectory() as tmp:
            registros = cargar_reportes(self._carpeta_de_prueba(tmp))
            resultados = buscar_items(registros, ["37014"], incluir_no_encontrados=False)
            restaurantes = {r.restaurante_id for r in resultados}
            self.assertNotIn("6004", restaurantes)
            self.assertIn("6001", restaurantes)


class PruebaDiccionario(unittest.TestCase):
    def test_carga_desde_excel(self):
        diccionario = DiccionarioRestaurantes.desde_excel(DICCIONARIO)
        self.assertGreaterEqual(len(diccionario), 20)
        self.assertEqual(diccionario.nombre("6001"), "Merida")
        self.assertEqual(diccionario.nombre("6013"), "Playa del Carmen")
        self.assertEqual(diccionario.nombre("6008"), "Pensiones")  # "6008 -Pensiones"
        self.assertEqual(diccionario.nombre("9999", "?"), "?")

    def test_localizar_en_carpeta(self):
        self.assertIsNotNone(DiccionarioRestaurantes.localizar(RAIZ))


class PruebaBusqueda(unittest.TestCase):
    def test_normalizar_entradas(self):
        self.assertEqual(normalizar_busquedas(["37014, 37021"]), ["37014", "37021"])
        self.assertEqual(normalizar_busquedas(["37014-37016"]), ["37014", "37015", "37016"])
        self.assertEqual(normalizar_busquedas(["37014, 37014"]), ["37014"])
        self.assertEqual(normalizar_busquedas([" alitas "]), ["alitas"])

    def test_busqueda_en_carpeta_con_varios_periodos(self):
        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            original = REPORTE_EJEMPLO.read_text(encoding="utf-8")
            (carpeta / "6001 (3-6).txt").write_text(original, encoding="utf-8")
            (carpeta / "6002 (10-13).txt").write_text(
                original.replace("#6001", "#6002").replace("8/3/26 to 8/6/26", "8/10/26 to 8/13/26"),
                encoding="utf-8",
            )
            self.assertEqual(len(listar_archivos(carpeta)), 2)

            registros = cargar_reportes(carpeta)
            diccionario = DiccionarioRestaurantes.desde_excel(DICCIONARIO)
            resultados = buscar_items(registros, ["37014", "99999"], diccionario)

            # La clave del periodo son las fechas que vienen DENTRO del reporte.
            ventas = {
                (r.restaurante_id, r.fecha_inicio): r
                for r in resultados
                if r.item_numero == 37014
            }
            primero = ventas[("6001", date(2026, 8, 3))]
            self.assertEqual(primero.unidades_vendidas, 98)
            self.assertEqual(primero.restaurante_nombre, "Merida")
            self.assertEqual(primero.fecha_fin, date(2026, 8, 6))
            segundo = ventas[("6002", date(2026, 8, 10))]
            self.assertEqual(segundo.restaurante_nombre, "Altabrisa")
            self.assertEqual(segundo.fecha_fin, date(2026, 8, 13))

            faltantes = [r for r in resultados if r.busqueda == "99999"]
            self.assertEqual(len(faltantes), 2)
            self.assertTrue(all(not r.encontrado and r.unidades_vendidas == 0 for r in faltantes))

    def test_las_fechas_vienen_del_reporte_no_del_nombre(self):
        """Si el nombre del archivo miente, mandan las fechas del reporte."""
        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            # El nombre dice 20-23, pero adentro el reporte es del 3 al 6.
            (carpeta / "6001 (20-23).txt").write_text(
                REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
            )
            registros = cargar_reportes(carpeta)
            (resultado,) = buscar_items(registros, ["37014"])
            self.assertEqual(resultado.fecha_inicio, date(2026, 8, 3))
            self.assertEqual(resultado.fecha_fin, date(2026, 8, 6))
            self.assertFalse(hasattr(resultado, "periodo"))
            self.assertIn("20-23", registros[0].aviso_periodo)

    def test_sin_aviso_cuando_el_nombre_coincide(self):
        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / REPORTE_EJEMPLO.name).write_text(
                REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
            )
            (registro,) = cargar_reportes(tmp)
            self.assertEqual(registro.aviso_periodo, "")

    def test_excluir_no_encontrados(self):
        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / REPORTE_EJEMPLO.name).write_text(
                REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
            )
            registros = cargar_reportes(tmp)
            resultados = buscar_items(registros, ["99999"], incluir_no_encontrados=False)
            self.assertEqual(resultados, [])


class PruebaExportacion(unittest.TestCase):
    def test_genera_las_hojas_fijas(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            (Path(tmp) / REPORTE_EJEMPLO.name).write_text(
                REPORTE_EJEMPLO.read_text(encoding="utf-8"), encoding="utf-8"
            )
            registros = cargar_reportes(tmp)
            resultados = buscar_items(registros, ["37014"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida", registros)
            self.assertTrue(ruta.exists())
            libro = load_workbook(ruta)
            self.assertEqual(
                libro.sheetnames,
                ["37014 03-08 a 06-08", "Resultados", "Resumen", "Archivos"],
            )
            hoja = libro["Resultados"]
            encabezados = [c.value for c in hoja[1]]
            self.assertEqual(
                encabezados[:6],
                [
                    "#ID Restaurante",
                    "Restaurante",
                    "# Item",
                    "Ventas (unidades)",
                    "Fecha inicio",
                    "Fecha fin",
                ],
            )
            # Ya no hay columna de "Periodo" en ninguna hoja.
            for nombre_hoja in libro.sheetnames:
                self.assertNotIn("Periodo", [c.value for c in libro[nombre_hoja][1]])
            self.assertEqual(hoja.cell(row=2, column=3).value, 37014)
            self.assertEqual(hoja.cell(row=2, column=4).value, 98)
            self.assertEqual(hoja.cell(row=2, column=5).value, "03/08/2026")
            self.assertEqual(hoja.cell(row=2, column=6).value, "06/08/2026")


class PruebaRutasConComillas(unittest.TestCase):
    """La ruta se puede pegar con o sin comillas (como la copia Windows)."""

    @staticmethod
    def _limpiar(texto: str) -> str:
        import importlib.util

        ruta = PROYECTO / "buscar_items.py"
        especificacion = importlib.util.spec_from_file_location("buscar_items", ruta)
        modulo = importlib.util.module_from_spec(especificacion)
        especificacion.loader.exec_module(modulo)
        return modulo._limpiar_ruta(texto)

    def test_quita_comillas(self):
        esperado = r"C:\Users\velascoh\Desktop\Hora Bostons"
        for entrada in (
            f'"{esperado}"',
            f"'{esperado}'",
            f"\u201c{esperado}\u201d",
            f'  "{esperado}"  ',
            esperado,
        ):
            with self.subTest(entrada=entrada):
                self.assertEqual(self._limpiar(entrada), esperado)

    def test_no_toca_una_ruta_normal(self):
        self.assertEqual(self._limpiar("  datos  "), "datos")


class PruebaHojasPorCombinacion(unittest.TestCase):
    """Una hoja por cada producto buscado y cada periodo encontrado."""

    def _carpeta_con_dos_periodos(self, tmp: str) -> Path:
        carpeta = Path(tmp)
        original = REPORTE_EJEMPLO.read_text(encoding="utf-8")
        (carpeta / "6001 (3-6).txt").write_text(original, encoding="utf-8")
        (carpeta / "6002 (10-13).txt").write_text(
            original.replace("#6001", "#6002").replace("8/3/26 to 8/6/26", "8/10/26 to 8/13/26"),
            encoding="utf-8",
        )
        return carpeta

    def test_dos_productos_por_dos_periodos_son_cuatro_hojas(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["37014", "37021"])
            self.assertEqual(contar_combinaciones(resultados), 4)

            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            libro = load_workbook(ruta)
            self.assertEqual(
                libro.sheetnames,
                [
                    "37014 03-08 a 06-08",
                    "37014 10-08 a 13-08",
                    "37021 03-08 a 06-08",
                    "37021 10-08 a 13-08",
                    "Resultados",
                    "Resumen",
                    "Archivos",
                ],
            )

            hoja = libro["37014 03-08 a 06-08"]
            self.assertIn("PLATO DE EQUIPO", hoja.cell(row=1, column=1).value)
            self.assertIn("03/08/2026", hoja.cell(row=1, column=1).value)
            self.assertEqual(hoja.cell(row=2, column=1).value, "#ID Restaurante")
            # Solo el restaurante de ese periodo, con sus 98 unidades.
            self.assertEqual(hoja.cell(row=3, column=1).value, "6001")
            self.assertEqual(hoja.cell(row=3, column=5).value, 98)
            # Filas 4-9: los 6 restaurantes de Oracle, sin ventas todavia.
            self.assertEqual(hoja.cell(row=4, column=1).value, "6016")
            self.assertIsNone(hoja.cell(row=4, column=5).value)
            self.assertEqual(hoja.cell(row=10, column=1).value, "TOTAL")
            self.assertEqual(hoja.cell(row=10, column=5).value, "=SUM(E3:E9)")

    def test_tres_productos_por_dos_periodos_son_seis_hojas(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["37014", "37021", "37031"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            hojas = load_workbook(ruta).sheetnames
            self.assertEqual(len(hojas), 6 + 3)  # 6 combinaciones + las 3 hojas fijas

    def test_columnas_solo_hasta_ventas_en_pesos(self):
        """Las hojas de producto x periodo llegan hasta 'Ventas ($)'."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["37014"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            libro = load_workbook(ruta)

            hoja = libro["37014 03-08 a 06-08"]
            encabezados = [c.value for c in hoja[2] if c.value]
            self.assertEqual(
                encabezados,
                [
                    "#ID Restaurante",
                    "Restaurante",
                    "# Item",
                    "Nombre del item",
                    "Ventas (unidades)",
                    "Ventas ($)",
                ],
            )
            # Las columnas G..M siguen estando en Resultados.
            self.assertIn("Costo total", [c.value for c in libro["Resultados"][1]])

    def test_los_renglones_de_oracle_no_llevan_formato_especial(self):
        """Se ven igual que los demas: sin relleno, sin bordes, sin aviso."""
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["37014"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            hoja = load_workbook(ruta)["37014 03-08 a 06-08"]

            etiquetas = [f[0] for f in hoja.iter_rows(values_only=True)]
            self.assertNotIn("Capturar a mano (Oracle)", etiquetas)

            fila_posi = hoja.cell(row=3, column=1)      # renglon leido del reporte
            fila_oracle = hoja.cell(row=4, column=1)    # primer renglon de Oracle
            self.assertEqual(fila_oracle.fill.fgColor.rgb, fila_posi.fill.fgColor.rgb)
            self.assertEqual(fila_oracle.font.b, fila_posi.font.b)
            self.assertEqual(fila_oracle.border.left.style, fila_posi.border.left.style)

    def test_renglones_de_oracle_solo_traen_el_restaurante(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["37014"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            hoja = load_workbook(ruta)["37014 03-08 a 06-08"]

            filas = list(hoja.iter_rows(values_only=True))
            # Van al final, justo antes del renglon de totales.
            oracle = filas[-1 - len(RESTAURANTES_ORACLE) : -1]

            # Salen ordenados por numero de restaurante.
            self.assertEqual(
                [(f[0], f[1]) for f in oracle],
                sorted(RESTAURANTES_ORACLE, key=lambda par: par[0]),
            )
            self.assertEqual([f[0] for f in oracle], sorted(f[0] for f in oracle))
            for fila in oracle:
                # En Oracle el item tiene otra numeracion y otro nombre: se
                # dejan vacios, igual que las ventas.
                self.assertIsNone(fila[2])
                self.assertIsNone(fila[3])
                self.assertIsNone(fila[4])
                self.assertIsNone(fila[5])

    def test_el_total_es_formula_e_incluye_los_de_oracle(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["37014"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            hoja = load_workbook(ruta)["37014 03-08 a 06-08"]

            fila_total = hoja.max_row
            self.assertEqual(hoja.cell(row=fila_total, column=1).value, "TOTAL")
            # El rango llega hasta el ultimo renglon de Oracle.
            self.assertEqual(
                hoja.cell(row=fila_total, column=5).value, f"=SUM(E3:E{fila_total - 1})"
            )
            self.assertEqual(
                hoja.cell(row=fila_total, column=6).value, f"=SUM(F3:F{fila_total - 1})"
            )

    def test_no_repite_un_restaurante_de_oracle_que_si_vino_en_los_datos(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = Path(tmp)
            original = REPORTE_EJEMPLO.read_text(encoding="utf-8")
            (carpeta / "6001 (3-6).txt").write_text(original, encoding="utf-8")
            (carpeta / "6016 (3-6).txt").write_text(
                original.replace("#6001", "#6016"), encoding="utf-8"
            )
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["37014"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            hoja = load_workbook(ruta)["37014 03-08 a 06-08"]

            identificadores = [f[0] for f in hoja.iter_rows(values_only=True)]
            self.assertEqual(identificadores.count("6016"), 1)

    def test_busqueda_por_nombre_junta_sus_items_en_una_hoja(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            resultados = buscar_items(registros, ["alitas"])
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            libro = load_workbook(ruta)
            hoja = libro["alitas 03-08 a 06-08"]
            self.assertIn("items", hoja.cell(row=1, column=1).value)
            self.assertGreater(hoja.max_row, 10)

    def test_nombres_de_hoja_validos_y_unicos(self):
        usados: set = set()
        largo = "un nombre de producto larguisimo que no cabe"
        primero = _nombre_de_hoja(largo, date(2026, 8, 3), date(2026, 8, 6), usados)
        segundo = _nombre_de_hoja(largo, date(2026, 8, 3), date(2026, 8, 6), usados)
        for nombre in (primero, segundo):
            self.assertLessEqual(len(nombre), 31)
            self.assertFalse(set(nombre) & set(r"[]:*?/\\"))
        self.assertNotEqual(primero, segundo)

        con_prohibidos = _nombre_de_hoja("pizza/queso*", None, None, usados)
        self.assertNotIn("/", con_prohibidos)
        self.assertIn("sin fechas", con_prohibidos)

    def test_tope_de_hojas(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            carpeta = self._carpeta_con_dos_periodos(tmp)
            registros = cargar_reportes(carpeta)
            # 100 numeros x 2 periodos = 200 combinaciones, por encima del tope.
            busquedas = [str(n) for n in range(37000, 37100)]
            resultados = buscar_items(registros, busquedas)
            self.assertEqual(contar_combinaciones(resultados), 200)
            ruta = exportar_resultados(resultados, Path(tmp) / "salida.xlsx", registros)
            hojas = load_workbook(ruta).sheetnames
            self.assertEqual(len(hojas), MAXIMO_HOJAS_COMBINACION + 3)
            # La hoja Resultados conserva todos los renglones.
            self.assertEqual(load_workbook(ruta)["Resultados"].max_row, len(resultados) + 1)


if __name__ == "__main__":
    unittest.main()

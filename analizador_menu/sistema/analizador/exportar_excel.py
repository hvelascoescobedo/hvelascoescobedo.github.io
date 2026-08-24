"""Generacion del archivo de Excel con los resultados de la busqueda.

El libro tiene estas hojas, en este orden:

* **Una hoja por cada combinacion de producto y periodo** (van al principio) --
  si se buscan 2 productos y hay 2 periodos, salen 4 hojas (producto 1 periodo
  1, producto 1 periodo 2, producto 2 periodo 1, producto 2 periodo 2). Cada
  una llega hasta la columna "Ventas ($)" e incluye, con el mismo formato que
  los demas, los renglones de los restaurantes de Oracle listos para capturarse
  a mano; el total va con formula para que se actualice al llenarlos.
* **Resultados** -- un renglon por item / restaurante / periodo, con todas las
  columnas. Las primeras son las pedidas: #ID de restaurante, nombre, numero de
  item y cuantas ventas (unidades vendidas).
* **Resumen**    -- total de unidades e importe por item y periodo.
* **Archivos**   -- registro de los archivos leidos.

Las fechas de inicio y fin de cada renglon salen del propio reporte, no del
nombre del archivo.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .busqueda import RegistroArchivo, ResultadoItem
from .configuracion import RESTAURANTES_ORACLE

RELLENO_ENCABEZADO = PatternFill("solid", fgColor="1F4E78")
FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF")

COLUMNAS_RESULTADOS = [
    ("#ID Restaurante", 16),
    ("Restaurante", 26),
    ("# Item", 10),
    ("Ventas (unidades)", 18),
    ("Fecha inicio", 14),
    ("Fecha fin", 14),
    ("Nombre del item", 24),
    ("Categoria", 22),
    ("Subcategoria", 24),
    ("Ventas ($)", 14),
    ("Precio de menu", 15),
    ("Costo total", 14),
    ("Utilidad", 14),
    ("Busqueda", 14),
    ("Encontrado", 12),
    ("Archivo", 26),
]

COLUMNAS_RESUMEN = [
    ("Busqueda", 14),
    ("# Item", 10),
    ("Nombre del item", 24),
    ("Fecha inicio", 14),
    ("Fecha fin", 14),
    ("Restaurantes con venta", 22),
    ("Ventas (unidades)", 18),
    ("Ventas ($)", 16),
]

# Hojas por combinacion producto x periodo (el producto y las fechas son fijos
# en cada hoja, por eso van en el titulo y no se repiten en cada renglon).
COLUMNAS_COMBINACION = [
    ("#ID Restaurante", 16),
    ("Restaurante", 26),
    ("# Item", 10),
    ("Nombre del item", 24),
    ("Ventas (unidades)", 18),
    ("Ventas ($)", 14),
]
COLUMNA_UNIDADES = 5
COLUMNA_IMPORTE = 6

# Tope de seguridad: Excel se vuelve inmanejable con miles de hojas.
MAXIMO_HOJAS_COMBINACION = 150

CARACTERES_PROHIBIDOS = set(r"[]:*?/\\")
LARGO_MAXIMO_HOJA = 31

COLUMNAS_ARCHIVOS = [
    ("Archivo", 30),
    ("#ID Restaurante", 16),
    ("Restaurante", 26),
    ("Fecha inicio", 14),
    ("Fecha fin", 14),
    ("Items leidos", 14),
    ("Estado", 30),
    ("Aviso", 46),
]


def _escribir_encabezado(hoja, columnas: Sequence[tuple[str, int]], fila: int = 1) -> None:
    for indice, (titulo, ancho) in enumerate(columnas, start=1):
        celda = hoja.cell(row=fila, column=indice, value=titulo)
        celda.fill = RELLENO_ENCABEZADO
        celda.font = FUENTE_ENCABEZADO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[get_column_letter(indice)].width = ancho
    hoja.freeze_panes = f"A{fila + 1}"
    hoja.auto_filter.ref = f"A{fila}:{get_column_letter(len(columnas))}{fila}"


def _formato_fecha(valor: date | None):
    return valor.strftime("%d/%m/%Y") if valor else ""


def _hoja_resultados(libro: Workbook, resultados: Sequence[ResultadoItem]) -> None:
    hoja = libro.create_sheet("Resultados")
    _escribir_encabezado(hoja, COLUMNAS_RESULTADOS)

    for fila, resultado in enumerate(resultados, start=2):
        valores = [
            resultado.restaurante_id,
            resultado.restaurante_nombre,
            resultado.item_numero if resultado.item_numero is not None else resultado.busqueda,
            resultado.unidades_vendidas,
            _formato_fecha(resultado.fecha_inicio),
            _formato_fecha(resultado.fecha_fin),
            resultado.item_nombre,
            resultado.categoria,
            resultado.subcategoria,
            round(resultado.ventas_importe, 2),
            round(resultado.precio_menu, 2),
            round(resultado.costo_total, 2),
            round(resultado.utilidad, 2),
            resultado.busqueda,
            "Si" if resultado.encontrado else "No",
            resultado.archivo,
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
        for columna in (10, 11, 12, 13):
            hoja.cell(row=fila, column=columna).number_format = "#,##0.00"

    hoja.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS_RESULTADOS))}{max(len(resultados) + 1, 1)}"


def _hoja_resumen(libro: Workbook, resultados: Sequence[ResultadoItem]) -> None:
    hoja = libro.create_sheet("Resumen")
    _escribir_encabezado(hoja, COLUMNAS_RESUMEN)

    acumulado: dict[tuple, dict] = defaultdict(
        lambda: {"nombre": "", "restaurantes": set(), "unidades": 0, "importe": 0.0}
    )
    for resultado in resultados:
        clave = (
            resultado.busqueda,
            resultado.item_numero,
            resultado.fecha_inicio,
            resultado.fecha_fin,
        )
        registro = acumulado[clave]
        if resultado.item_nombre and not registro["nombre"]:
            registro["nombre"] = resultado.item_nombre
        if resultado.unidades_vendidas:
            registro["restaurantes"].add(resultado.restaurante_id)
        registro["unidades"] += resultado.unidades_vendidas
        registro["importe"] += resultado.ventas_importe

    fila = 2
    for (busqueda, numero, inicio, fin), registro in sorted(
        acumulado.items(),
        key=lambda par: (str(par[0][0]), str(par[0][1]), par[0][2] or date.min),
    ):
        valores = [
            busqueda,
            numero if numero is not None else "",
            registro["nombre"],
            _formato_fecha(inicio),
            _formato_fecha(fin),
            len(registro["restaurantes"]),
            registro["unidades"],
            round(registro["importe"], 2),
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
        hoja.cell(row=fila, column=8).number_format = "#,##0.00"
        fila += 1


# ---------------------------------------------------------------------------
# Una hoja por cada combinacion de producto y periodo
# ---------------------------------------------------------------------------


def _etiqueta_periodo(inicio: date | None, fin: date | None) -> str:
    """Periodo corto para el nombre de la hoja: '03-08 a 06-08'."""
    if inicio and fin:
        return f"{inicio:%d-%m} a {fin:%d-%m}"
    return "sin fechas"


def _nombre_de_hoja(busqueda: str, inicio: date | None, fin: date | None, usados: set) -> str:
    """Nombre valido y unico para la hoja (Excel: 31 caracteres, sin []:*?/\\)."""
    base = f"{busqueda} {_etiqueta_periodo(inicio, fin)}"
    limpio = "".join(" " if c in CARACTERES_PROHIBIDOS else c for c in base).strip()
    nombre = limpio[:LARGO_MAXIMO_HOJA] or "Hoja"

    if nombre.casefold() in usados:
        # Se recorta para dejar lugar al sufijo " (2)", " (3)", ...
        for intento in range(2, 1000):
            sufijo = f" ({intento})"
            candidato = f"{nombre[:LARGO_MAXIMO_HOJA - len(sufijo)].strip()}{sufijo}"
            if candidato.casefold() not in usados:
                nombre = candidato
                break
    usados.add(nombre.casefold())
    return nombre


def _renglones_oracle(
    hoja,
    fila: int,
    renglones: Sequence[ResultadoItem],
    busqueda: str,
) -> int:
    """Deja los renglones de Oracle rotulados y vacios para capturar a mano.

    Van con el mismo formato que los demas renglones; solo quedan vacias las
    celdas de ventas.  Devuelve la fila siguiente a los renglones escritos.
    """
    ya_presentes = {r.restaurante_id for r in renglones}
    pendientes = [(i, n) for i, n in RESTAURANTES_ORACLE if i not in ya_presentes]
    if not pendientes:
        return fila

    # Si se busco un numero de item se rotula tambien el item; si se busco por
    # nombre puede haber varios, asi que esas celdas se dejan vacias.
    numeros = {r.item_numero for r in renglones if r.item_numero is not None}
    numero_item = next(iter(numeros)) if len(numeros) == 1 else None
    if numero_item is None and busqueda.isdigit():
        numero_item = int(busqueda)
    nombre_item = next((r.item_nombre for r in renglones if r.item_nombre), "")
    if len(numeros) > 1:
        nombre_item = ""

    for identificador, nombre in pendientes:
        hoja.cell(row=fila, column=1, value=identificador)
        hoja.cell(row=fila, column=2, value=nombre)
        if numero_item is not None:
            hoja.cell(row=fila, column=3, value=numero_item)
        if nombre_item:
            hoja.cell(row=fila, column=4, value=nombre_item)
        hoja.cell(row=fila, column=COLUMNA_IMPORTE).number_format = "#,##0.00"
        fila += 1

    return fila


def _hoja_de_combinacion(
    libro: Workbook,
    busqueda: str,
    inicio: date | None,
    fin: date | None,
    renglones: Sequence[ResultadoItem],
    usados: set,
) -> None:
    """Escribe la hoja de un producto en un periodo, con su total al final."""
    hoja = libro.create_sheet(_nombre_de_hoja(busqueda, inicio, fin, usados))

    nombre_item = next((r.item_nombre for r in renglones if r.item_nombre), "")
    numeros = sorted({r.item_numero for r in renglones if r.item_numero is not None})
    if len(numeros) == 1:
        descripcion = f"Item {numeros[0]}" + (f" - {nombre_item}" if nombre_item else "")
    else:
        descripcion = f"Busqueda '{busqueda}' ({len(numeros)} items)"
    periodo = (
        f"del {inicio:%d/%m/%Y} al {fin:%d/%m/%Y}"
        if inicio and fin
        else "(el reporte no traia fechas)"
    )

    titulo = hoja.cell(row=1, column=1, value=f"{descripcion}  |  {periodo}")
    titulo.font = Font(bold=True, size=12)
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS_COMBINACION))

    _escribir_encabezado(hoja, COLUMNAS_COMBINACION, fila=2)

    primera_fila = 3
    fila = primera_fila
    for renglon in sorted(renglones, key=lambda r: (r.restaurante_id, r.item_numero or 0)):
        valores = [
            renglon.restaurante_id,
            renglon.restaurante_nombre,
            renglon.item_numero if renglon.item_numero is not None else renglon.busqueda,
            renglon.item_nombre,
            renglon.unidades_vendidas,
            round(renglon.ventas_importe, 2),
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
        hoja.cell(row=fila, column=COLUMNA_IMPORTE).number_format = "#,##0.00"
        fila += 1

    fila = _renglones_oracle(hoja, fila, renglones, busqueda)
    ultima_fila = fila - 1

    # El total va con formula para que se actualice al capturar los de Oracle.
    hoja.cell(row=fila, column=1, value="TOTAL").font = Font(bold=True)
    for columna in (COLUMNA_UNIDADES, COLUMNA_IMPORTE):
        letra = get_column_letter(columna)
        celda = hoja.cell(row=fila, column=columna)
        celda.value = f"=SUM({letra}{primera_fila}:{letra}{ultima_fila})"
        celda.font = Font(bold=True)
        if columna == COLUMNA_IMPORTE:
            celda.number_format = "#,##0.00"


def _hojas_por_combinacion(libro: Workbook, resultados: Sequence[ResultadoItem]) -> int:
    """Crea una hoja por cada producto buscado y cada periodo encontrado."""
    combinaciones: dict[tuple, list[ResultadoItem]] = defaultdict(list)
    for resultado in resultados:
        combinaciones[(resultado.busqueda, resultado.fecha_inicio, resultado.fecha_fin)].append(
            resultado
        )

    orden_busquedas = {}
    for resultado in resultados:  # se respeta el orden en que se pidieron
        orden_busquedas.setdefault(resultado.busqueda, len(orden_busquedas))

    claves = sorted(
        combinaciones,
        key=lambda clave: (orden_busquedas[clave[0]], clave[1] or date.min),
    )

    usados: set = set()
    for numero, clave in enumerate(claves, start=1):
        if numero > MAXIMO_HOJAS_COMBINACION:
            break
        busqueda, inicio, fin = clave
        _hoja_de_combinacion(libro, busqueda, inicio, fin, combinaciones[clave], usados)

    return len(claves)


def _hoja_archivos(
    libro: Workbook,
    registros: Sequence[RegistroArchivo],
    resultados: Sequence[ResultadoItem],
) -> None:
    hoja = libro.create_sheet("Archivos")
    _escribir_encabezado(hoja, COLUMNAS_ARCHIVOS)

    nombres = {
        resultado.restaurante_id: resultado.restaurante_nombre
        for resultado in resultados
        if resultado.restaurante_nombre
    }

    for fila, registro in enumerate(registros, start=2):
        if registro.reporte is None:
            estado = f"ERROR: {registro.error}"
        elif not registro.reporte.items:
            estado = "SIN ITEMS (revisar el formato del archivo)"
        else:
            estado = "OK"
        valores = [
            registro.info.nombre_archivo,
            registro.restaurante_id,
            nombres.get(registro.restaurante_id, ""),
            _formato_fecha(registro.fecha_inicio),
            _formato_fecha(registro.fecha_fin),
            len(registro.reporte.items) if registro.reporte else 0,
            estado,
            registro.aviso_periodo,
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)


def exportar_resultados(
    resultados: Sequence[ResultadoItem],
    ruta_salida: str | Path,
    registros: Sequence[RegistroArchivo] = (),
) -> Path:
    """Escribe el Excel de resultados y devuelve la ruta final."""
    ruta_salida = Path(ruta_salida)
    if ruta_salida.suffix.lower() != ".xlsx":
        ruta_salida = ruta_salida.with_suffix(".xlsx")
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    # Si el archivo esta abierto en Excel no se puede sobrescribir: se versiona.
    if ruta_salida.exists():
        try:
            with open(ruta_salida, "ab"):
                pass
        except OSError:
            marca = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_salida = ruta_salida.with_name(f"{ruta_salida.stem}_{marca}.xlsx")

    libro = Workbook()
    hoja_vacia = libro.active  # la que crea openpyxl por defecto; se quita al final

    # Primero las hojas de producto x periodo y despues las hojas generales.
    _hojas_por_combinacion(libro, resultados)
    _hoja_resultados(libro, resultados)
    _hoja_resumen(libro, resultados)
    if registros:
        _hoja_archivos(libro, registros, resultados)

    libro.remove(hoja_vacia)
    libro.save(ruta_salida)
    return ruta_salida


def contar_combinaciones(resultados: Sequence[ResultadoItem]) -> int:
    """Cuantas hojas de producto x periodo corresponden a estos resultados."""
    return len({(r.busqueda, r.fecha_inicio, r.fecha_fin) for r in resultados})

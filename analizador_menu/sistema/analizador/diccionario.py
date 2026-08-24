"""Catalogo de restaurantes: #ID -> nombre.

Lee el Excel 'Diccionario_restaurantes.xlsx', que puede venir en dos formas:

1. Una sola columna con el texto completo ("6001 - Merida"), de donde se
   separan los primeros 4 digitos como ID y el resto como nombre.
2. Dos columnas separadas (ID y nombre), en cualquier orden.
"""

from __future__ import annotations

import re
from pathlib import Path

try:  # openpyxl es obligatorio para leer/escribir Excel
    from openpyxl import load_workbook
except ImportError as error:  # pragma: no cover - depende del entorno
    raise ImportError(
        "Falta la libreria 'openpyxl'. Instalala con: pip install -r requirements.txt"
    ) from error

RE_ID = re.compile(r"(?P<id>\d{4})")
NOMBRES_ARCHIVO = (
    "Diccionario_restaurantes.xlsx",
    "diccionario_restaurantes.xlsx",
    "Diccionario restaurantes.xlsx",
)


class DiccionarioRestaurantes:
    """Traduce el numero de restaurante a su nombre."""

    def __init__(self, mapa: dict[str, str] | None = None) -> None:
        self._mapa: dict[str, str] = dict(mapa or {})

    # -- construccion --------------------------------------------------
    @classmethod
    def desde_excel(cls, ruta: str | Path) -> "DiccionarioRestaurantes":
        """Carga el diccionario desde un archivo .xlsx."""
        ruta = Path(ruta)
        libro = load_workbook(ruta, data_only=True, read_only=True)
        mapa: dict[str, str] = {}
        try:
            for hoja in libro.worksheets:
                for fila in hoja.iter_rows(values_only=True):
                    identificador, nombre = cls._interpretar_fila(fila)
                    if identificador and identificador not in mapa:
                        mapa[identificador] = nombre
        finally:
            libro.close()
        return cls(mapa)

    @classmethod
    def localizar(cls, carpeta: str | Path) -> "DiccionarioRestaurantes | None":
        """Busca el diccionario dentro de una carpeta (o sus subcarpetas)."""
        carpeta = Path(carpeta)
        for nombre in NOMBRES_ARCHIVO:
            candidato = carpeta / nombre
            if candidato.exists():
                return cls.desde_excel(candidato)
        for candidato in sorted(carpeta.glob("**/*iccionario*.xlsx")):
            if not candidato.name.startswith("~$"):
                return cls.desde_excel(candidato)
        return None

    @staticmethod
    def _interpretar_fila(fila: tuple) -> tuple[str | None, str]:
        """Extrae (id, nombre) de una fila del Excel; ignora encabezados."""
        celdas = [str(c).strip() for c in fila if c is not None and str(c).strip()]
        if not celdas:
            return None, ""
        # Se ignoran encabezados y formulas sin valor calculado.
        celdas = [c for c in celdas if not c.startswith("=")]
        if not celdas:
            return None, ""

        texto = celdas[0]
        if texto.upper() in {"RESTAURANT", "RESTAURANTE", "CODIGO", "CÓDIGO", "ID"}:
            return None, ""

        coincidencia = RE_ID.search(texto)
        if not coincidencia:
            return None, ""
        identificador = coincidencia.group("id")

        # Nombre: lo que sigue al ID en la misma celda, o la segunda celda.
        resto = texto[coincidencia.end():].strip(" -–—:\t")
        if not resto and len(celdas) > 1:
            segunda = celdas[1]
            resto = segunda[RE_ID.search(segunda).end():].strip(" -–—:\t") if RE_ID.search(segunda) else segunda
        return identificador, resto or texto

    # -- consulta ------------------------------------------------------
    def nombre(self, restaurante_id: str | int | None, por_defecto: str = "") -> str:
        """Nombre del restaurante; cadena vacia (o ``por_defecto``) si no esta."""
        if restaurante_id is None:
            return por_defecto
        return self._mapa.get(str(restaurante_id).strip(), por_defecto)

    def __contains__(self, restaurante_id: object) -> bool:
        return str(restaurante_id) in self._mapa

    def __len__(self) -> int:
        return len(self._mapa)

    def items(self):
        return self._mapa.items()
